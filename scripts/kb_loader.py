from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from schemas import KnowledgeBaseEntry, normalize_field_name


DEFAULT_ALIAS_MAP: dict[str, list[str]] = {
    normalize_field_name("供应商名称"): ["磋商供应商名称", "磋商单位名称", "单位名称", "乙方", "企业名称"],
    normalize_field_name("负责人姓名"): ["法定代表人", "负责人姓名", "法定代表人/负责人", "姓名"],
    normalize_field_name("地址"): ["通讯地址", "注册地址", "项目地址"],
    normalize_field_name("电话"): ["联系电话"],
    normalize_field_name("开户银行"): ["银行名称"],
    normalize_field_name("账号"): ["银行账号", "开户账号"],
    normalize_field_name("统一社会信用代码"): ["法人证书号", "统一社会信用代码；事业单位（法人证书号）；团体组织（法人证书号）"],
    normalize_field_name("网址"): ["网站"],
    normalize_field_name("组织结构"): ["组织机构"],
    normalize_field_name("负责人职务"): ["职务", "职务名称", "法定代表人职务", "单位负责人职务"],
}


class KnowledgeBase:
    def __init__(self, entries: dict[str, KnowledgeBaseEntry]):
        self.entries = entries
        self.alias_to_key: dict[str, str] = {}
        for key, entry in entries.items():
            self.alias_to_key[key] = key
            for alias in entry.aliases:
                self.alias_to_key[normalize_field_name(alias)] = key

    def lookup(self, field_name: str) -> KnowledgeBaseEntry | None:
        normalized = normalize_field_name(field_name)
        key = self.alias_to_key.get(normalized)
        return self.entries.get(key) if key else None

    def fuzzy_lookup(self, field_name: str) -> KnowledgeBaseEntry | None:
        normalized = normalize_field_name(field_name)
        best_key = ""
        best_score = 0
        for key, entry in self.entries.items():
            candidates = [key, *(normalize_field_name(alias) for alias in entry.aliases)]
            for candidate in candidates:
                if not candidate:
                    continue
                if candidate in normalized or normalized in candidate:
                    score = min(len(candidate), len(normalized))
                    if score > best_score:
                        best_key = key
                        best_score = score
        return self.entries.get(best_key) if best_key else None

    def has_field(self, field_name: str) -> bool:
        return self.lookup(field_name) is not None or self.fuzzy_lookup(field_name) is not None

    def to_dict(self) -> dict[str, dict]:
        return {key: asdict(value) for key, value in self.entries.items()}


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_columns(header_row: list[str]) -> tuple[int, int]:
    normalized_headers = [normalize_field_name(cell) for cell in header_row]
    field_col = -1
    value_col = -1
    for index, header in enumerate(normalized_headers):
        if field_col == -1 and header in {
            normalize_field_name("具体填写项"),
            normalize_field_name("字段名"),
            normalize_field_name("项目"),
        }:
            field_col = index
        if value_col == -1 and header in {
            normalize_field_name("备注"),
            normalize_field_name("字段值"),
            normalize_field_name("值"),
            normalize_field_name("内容"),
        }:
            value_col = index
    if field_col == -1 or value_col == -1:
        raise ValueError("Knowledge base header must contain field and value columns.")
    return field_col, value_col


def _normalize_kb_value(field_name: str, value: str) -> str:
    if normalize_field_name(field_name) != normalize_field_name("成立时间"):
        return value
    text = (value or "").strip()
    if not text:
        return text
    if re.fullmatch(r"\d+(?:\.0)?", text):
        try:
            parsed = from_excel(float(text))
        except (TypeError, ValueError):
            return text
        return f"{parsed.year}年{parsed.month}月"
    return text


def load_knowledge_base(xlsx_path: str) -> KnowledgeBase:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Knowledge base file does not exist: {xlsx_path}")

    workbook = load_workbook(path, data_only=True)
    target_sheet = workbook.active
    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Knowledge base worksheet is empty.")

    header = [_clean_cell(cell) for cell in rows[0]]
    field_col, value_col = _find_columns(header)
    entries: dict[str, KnowledgeBaseEntry] = {}

    for row in rows[1:]:
        cells = [_clean_cell(cell) for cell in row]
        if field_col >= len(cells):
            continue
        field_name = cells[field_col]
        if not field_name:
            continue
        value = cells[value_col] if value_col < len(cells) else ""
        value = _normalize_kb_value(field_name, value)
        normalized = normalize_field_name(field_name)
        aliases = DEFAULT_ALIAS_MAP.get(normalized, []).copy()
        entries[normalized] = KnowledgeBaseEntry(
            field_name=field_name,
            value=value,
            source_sheet=target_sheet.title,
            aliases=aliases,
        )

    if not entries:
        raise ValueError("Knowledge base did not yield any valid entries.")
    return KnowledgeBase(entries)
